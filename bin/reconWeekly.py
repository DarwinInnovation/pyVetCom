#!/usr/bin/python
# -*- coding: iso-8859-15 -*-

__author__ = 'richardm'

import logging
import os
import PyVetCom
from datetime import date, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

xlsxfile = 'weekly_recon.xlsx'

side_thin = Side(border_style='thin')
side_thick = Side(border_style='thick')
style_default_top = Style(font=Font(size=9), border=Border(left=side_thin, right=side_thin, top=side_thick, bottom=side_thin))
style_default_bottom = Style(font=Font(size=9), border=Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thick))

def setCell(ws, col, row, val, style):
    cell = ws["%s%d"%(get_column_letter(col+0),row)]
    cell.value = val
    cell.style = style

def copyFormat(from_ws, to_ws):
    for col in range(1,20):
        for row in range(1,17):
            from_cell = from_ws["%s%d"%(get_column_letter(col),row)]
            to_cell = to_ws["%s%d"%(get_column_letter(col),row)]
            to_cell.value = from_cell.value
            to_cell.style = from_cell.style

        from_col = from_ws.column_dimensions["%s"%get_column_letter(col+0)]
        to_col = to_ws.column_dimensions["%s"%get_column_letter(col+0)]

        to_col.width = from_col.width

    for row in range(1,17):
        from_row = from_ws.row_dimensions[row]
        to_row = to_ws.row_dimensions[row]

        to_row.height = from_row.height


def setPrintOptions(ws):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = True

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25

if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)

    vc = PyVetCom.PyVetCom("192.168.3.200")

    start = date.today() - timedelta(weeks=1)
    while start.weekday() <> 0:
        start = start - timedelta(days=1)
    end = start + timedelta(days=6)

    dailyfigures = []

    day = start
    while day < end:
        wd = day.weekday()

        dailyfigures.insert(wd, [0,0,0,0,0,0])
        figs = dailyfigures[wd]
        for type in range(2, 6):
            sum = vc.SLDocs().ondate("DATETIME", day).eq("TYPE", type).sum('TOTAL')
            figs.insert(type, sum)
            print "%s: %d %f"%(day, type, sum)

        day += timedelta(days=1)

    if (os.path.exists(xlsxfile)):
        wb = load_workbook(filename = xlsxfile)
    else:
        wb = Workbook()

    title = start.strftime("%d-%m")
    template_ws = wb.get_sheet_by_name("Template")
    ws = wb.get_sheet_by_name(title)
    if ws is None:
        ws = wb.create_sheet()

    print "Title: %s"%title
    ws.title = (title)

    copyFormat(template_ws, ws)

    day = start
    while day < end:
        wd = day.weekday()
        row = (2*wd)+3

        cell = ws["A%d"%row]
        cell.value = day

        ws["N%d"%(row+1)] = ""

        figs = dailyfigures[wd]
        total = 0.0
        for type in range(2, 6):
            col = (type*3)-4
            cell =  ws["%s%d"%(get_column_letter(col), row)]
            cell.value = figs[type]
            total += figs[type]
        cell =  ws["N%d"%(row)]
        cell.value = total
        day += timedelta(days=1)

    setPrintOptions(ws)

    wb.save(filename=xlsxfile)
